# ========== main.py (COMPLETO MEJORADO) ==========
import json
import logging
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
from functools import wraps


from fastapi import FastAPI, APIRouter, Depends, HTTPException, status, Request
from fastapi.security import OAuth2PasswordRequestForm
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from pydantic import BaseModel, Field, validator
from models.schemas import GuardarSolucionRequest

from database import get_db, engine
from models.db_models import (
    Base, Facultad, Programa, Docente, Aula,
    CursoDB, HorarioGenerado, Usuario, Matricula,
    SolicitudRetiro, SolicitudCambio
)
from models.schemas import HorarioRequest, HorarioResponse
from csp_engine import CSPEngine
from csp_engine_profesional import CSPEngineProfesional
from auth import (
    authenticate_user, create_access_token, get_current_active_user,
    require_role, get_password_hash, ACCESS_TOKEN_EXPIRE_MINUTES,
    ForgotPasswordRequest, ResetPasswordRequest,
    generar_token_recuperacion, verificar_token_recuperacion,
    reset_tokens
)

import io
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from fastapi.responses import StreamingResponse

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Crear tablas al inicio
Base.metadata.create_all(bind=engine)

# ========== SCHEMAS MEJORADOS ==========
class FacultadCreate(BaseModel):
    nombre: str = Field(..., min_length=1, max_length=100)
    codigo: str = Field(..., min_length=1, max_length=10)
    
    @validator('codigo')
    def codigo_upper(cls, v):
        return v.upper()

class ProgramaCreate(BaseModel):
    nombre: str = Field(..., min_length=1, max_length=100)
    codigo: str = Field(..., min_length=1, max_length=10)
    facultad_id: int = Field(..., gt=0)
    
    @validator('codigo')
    def codigo_upper(cls, v):
        return v.upper()

class DocenteCreate(BaseModel):
    codigo: str = Field(..., min_length=1, max_length=20)
    nombre: str = Field(..., min_length=1)
    apellido: str = Field(..., min_length=1)
    email: str = Field(..., pattern=r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')
    disponibilidad: List[str] = []
    max_horas_semanales: int = Field(20, ge=4, le=40)
    
    @validator('codigo')
    def codigo_upper(cls, v):
        return v.upper()

class AulaCreate(BaseModel):
    codigo: str = Field(..., min_length=1, max_length=20)
    nombre: str = Field(..., min_length=1)
    capacidad: int = Field(..., ge=5, le=500)
    tipo: str = Field("teoria", pattern=r'^(teoria|laboratorio|taller|auditorio)$')
    edificio: str = "Principal"
    piso: int = Field(1, ge=0, le=20)
    recursos: List[str] = []
    
    @validator('codigo')
    def codigo_upper(cls, v):
        return v.upper()

class CursoCreate(BaseModel):
    codigo: str = Field(..., pattern=r'^[A-Z]{2,4}[0-9]{3,4}$')
    nombre: str = Field(..., min_length=1)
    creditos: int = Field(..., ge=1, le=6)
    semestre: int = Field(..., ge=1, le=12)
    programa_id: int = Field(..., gt=0)
    docente_id: int = Field(..., gt=0)
    max_estudiantes: int = Field(30, ge=5, le=200)
    tipo: str = Field("teoria", pattern=r'^(teoria|laboratorio|taller|practica)$')
    horas_semanales: int = Field(4, ge=2, le=20)
    
    @validator('codigo')
    def codigo_upper(cls, v):
        return v.upper()

class MatriculaInscripcion(BaseModel):
    horario_ids: List[int] = Field(..., min_length=1, max_length=10)
    
    @validator('horario_ids')
    def no_duplicados(cls, v):
        if len(v) != len(set(v)):
            raise ValueError('No se permiten IDs duplicados')
        return v

class UsuarioCreate(BaseModel):
    email: str
    password: str
    nombre: str
    apellido: str
    rol: str = Field("estudiante", pattern=r'^(admin|coordinador|docente|estudiante)$')

class HorariosAvanzadoRequest(BaseModel):
    semestre: int = Field(..., ge=1, le=12, description="Semestre a planificar")
    num_soluciones: int = Field(5, ge=1, le=10, description="Cantidad de alternativas")
    turno_preferido: Optional[str] = Field(
        None,
        pattern=r'^(mañana|tarde|noche)$',
        description="Turno preferido del estudiante (opcional)"
    )

# ========== FUNCIONES AUXILIARES ==========
def parse_json_field(value: str) -> List:
    """Helper para parsear campos JSON de la BD"""
    if value is None or value == '':
        return []
    try:
        return json.loads(value)
    except:
        return []

def serialize_json_field(value: List) -> str:
    """Helper para serializar listas a JSON"""
    if value is None:
        return '[]'
    return json.dumps(value)

def get_curso_nombre(db: Session, curso_id: int) -> str:
    """Obtener nombre del curso por ID"""
    curso = db.query(CursoDB).filter(CursoDB.id == curso_id).first()
    return curso.nombre if curso else "Desconocido"

def get_docente_nombre(db: Session, docente_id: int) -> str:
    """Obtener nombre completo del docente"""
    docente = db.query(Docente).filter(Docente.id == docente_id).first()
    return f"{docente.nombre} {docente.apellido}" if docente else "Desconocido"

# ========== INICIALIZAR APP ==========
app = FastAPI(title="OptiAcademic API", version="1.0.0")

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

router = APIRouter(prefix="/api/v1", tags=["OptiAcademic"])

# ========== ENDPOINTS DE CATÁLOGO ==========
@router.post("/facultades", status_code=status.HTTP_201_CREATED)
def crear_facultad(data: FacultadCreate, db: Session = Depends(get_db)):
    """Crear una nueva facultad"""
    # Verificar si ya existe
    existing = db.query(Facultad).filter(Facultad.codigo == data.codigo).first()
    if existing:
        raise HTTPException(400, f"Ya existe una facultad con código {data.codigo}")
    
    facultad = Facultad(**data.model_dump())
    db.add(facultad)
    db.commit()
    db.refresh(facultad)
    logger.info(f"Facultad creada: {facultad.codigo} - {facultad.nombre}")
    return facultad

@router.get("/facultades")
def listar_facultades(db: Session = Depends(get_db)):
    """Listar todas las facultades"""
    return db.query(Facultad).filter(Facultad.activo == True).all()

@router.post("/programas", status_code=status.HTTP_201_CREATED)
def crear_programa(data: ProgramaCreate, db: Session = Depends(get_db)):
    """Crear un nuevo programa académico"""
    # Verificar facultad
    facultad = db.query(Facultad).filter(Facultad.id == data.facultad_id).first()
    if not facultad:
        raise HTTPException(404, "Facultad no encontrada")
    
    existing = db.query(Programa).filter(Programa.codigo == data.codigo).first()
    if existing:
        raise HTTPException(400, f"Ya existe un programa con código {data.codigo}")
    
    programa = Programa(**data.model_dump())
    db.add(programa)
    db.commit()
    db.refresh(programa)
    logger.info(f"Programa creado: {programa.codigo} - {programa.nombre}")
    return programa

@router.get("/programas")
def listar_programas(facultad_id: Optional[int] = None, db: Session = Depends(get_db)):
    """Listar programas, opcionalmente filtrando por facultad"""
    query = db.query(Programa).filter(Programa.activo == True)
    if facultad_id:
        query = query.filter(Programa.facultad_id == facultad_id)
    return query.all()

@router.post("/docentes", status_code=status.HTTP_201_CREATED)
def crear_docente(data: DocenteCreate, db: Session = Depends(get_db)):
    """Crear un nuevo docente"""
    existing = db.query(Docente).filter(Docente.codigo == data.codigo).first()
    if existing:
        raise HTTPException(400, f"Ya existe un docente con código {data.codigo}")
    
    docente_data = data.model_dump()
    docente_data['disponibilidad'] = serialize_json_field(docente_data['disponibilidad'])
    
    docente = Docente(**docente_data)
    db.add(docente)
    db.commit()
    db.refresh(docente)
    
    return {
        "id": docente.id,
        "codigo": docente.codigo,
        "nombre": docente.nombre,
        "apellido": docente.apellido,
        "email": docente.email,
        "disponibilidad": parse_json_field(docente.disponibilidad),
        "max_horas_semanales": docente.max_horas_semanales,
        "activo": docente.activo
    }

@router.get("/docentes")
def listar_docentes(activo: Optional[bool] = True, db: Session = Depends(get_db)):
    """Listar docentes"""
    query = db.query(Docente)
    if activo is not None:
        query = query.filter(Docente.activo == activo)
    
    return [
        {
            "id": d.id,
            "codigo": d.codigo,
            "nombre": d.nombre,
            "apellido": d.apellido,
            "email": d.email,
            "disponibilidad": parse_json_field(d.disponibilidad),
            "max_horas_semanales": d.max_horas_semanales,
            "activo": d.activo
        }
        for d in query.all()
    ]

@router.post("/aulas", status_code=status.HTTP_201_CREATED)
def crear_aula(data: AulaCreate, db: Session = Depends(get_db)):
    """Crear una nueva aula"""
    existing = db.query(Aula).filter(Aula.codigo == data.codigo).first()
    if existing:
        raise HTTPException(400, f"Ya existe un aula con código {data.codigo}")
    
    aula_data = data.model_dump()
    aula_data['recursos'] = serialize_json_field(aula_data['recursos'])
    
    aula = Aula(**aula_data)
    db.add(aula)
    db.commit()
    db.refresh(aula)
    
    return {
        "id": aula.id,
        "codigo": aula.codigo,
        "nombre": aula.nombre,
        "capacidad": aula.capacidad,
        "tipo": aula.tipo,
        "edificio": aula.edificio,
        "piso": aula.piso,
        "recursos": parse_json_field(aula.recursos),
        "activa": aula.activa
    }

@router.get("/aulas")
def listar_aulas(activa: Optional[bool] = True, tipo: Optional[str] = None, db: Session = Depends(get_db)):
    """Listar aulas, opcionalmente filtrando por tipo y estado"""
    query = db.query(Aula)
    if activa is not None:
        query = query.filter(Aula.activa == activa)
    if tipo:
        query = query.filter(Aula.tipo == tipo)
    
    return [
        {
            "id": a.id,
            "codigo": a.codigo,
            "nombre": a.nombre,
            "capacidad": a.capacidad,
            "tipo": a.tipo,
            "edificio": a.edificio,
            "piso": a.piso,
            "recursos": parse_json_field(a.recursos),
            "activa": a.activa
        }
        for a in query.all()
    ]

# ========== CURSOS ==========

@router.get("/cursos")
def listar_cursos(
    semestre: Optional[int] = None,
    programa_id: Optional[int] = None,
    docente_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """Listar cursos con filtros opcionales (público, sin autenticación)"""
    query = db.query(CursoDB)
    if semestre:
        query = query.filter(CursoDB.semestre == semestre)
    if programa_id:
        query = query.filter(CursoDB.programa_id == programa_id)
    if docente_id:
        query = query.filter(CursoDB.docente_id == docente_id)

    return [
        {
            "id": c.id,
            "codigo": c.codigo,
            "nombre": c.nombre,
            "creditos": c.creditos,
            "semestre": c.semestre,
            "programa_id": c.programa_id,
            "docente_id": c.docente_id,
            "max_estudiantes": c.max_estudiantes,
            "tipo": c.tipo,
            "horas_semanales": c.horas_semanales,
            "horas_teoria": getattr(c, "horas_teoria", None),
            "horas_laboratorio": getattr(c, "horas_laboratorio", None),
            "horas_practica": getattr(c, "horas_practica", None),
            "minutos_por_hora": getattr(c, "minutos_por_hora", None),
        }
        for c in query.all()
    ]


@router.get("/cursos/{curso_id}")
def obtener_curso(curso_id: int, db: Session = Depends(get_db)):
    """Obtener un curso por ID"""
    curso = db.query(CursoDB).filter(CursoDB.id == curso_id).first()
    if not curso:
        raise HTTPException(404, "Curso no encontrado")
    return {
        "id": curso.id,
        "codigo": curso.codigo,
        "nombre": curso.nombre,
        "creditos": curso.creditos,
        "semestre": curso.semestre,
        "programa_id": curso.programa_id,
        "docente_id": curso.docente_id,
        "max_estudiantes": curso.max_estudiantes,
        "tipo": curso.tipo,
        "horas_semanales": curso.horas_semanales,
    }


@router.post("/cursos", status_code=status.HTTP_201_CREATED)
def crear_curso(
    curso: CursoCreate,
    current_user: Usuario = Depends(require_role(["admin", "coordinador"])),
    db: Session = Depends(get_db)
):
    """Crear un nuevo curso (solo admin/coordinador)"""
    existing = db.query(CursoDB).filter(CursoDB.codigo == curso.codigo).first()
    if existing:
        raise HTTPException(400, f"Ya existe un curso con código {curso.codigo}")

    nuevo_curso = CursoDB(**curso.model_dump())
    db.add(nuevo_curso)
    db.commit()
    db.refresh(nuevo_curso)
    logger.info(f"Curso creado: {nuevo_curso.codigo} - {nuevo_curso.nombre}")
    return nuevo_curso


@router.put("/cursos/{curso_id}")
def actualizar_curso(
    curso_id: int,
    curso_data: dict,
    current_user: Usuario = Depends(require_role(["admin", "coordinador"])),
    db: Session = Depends(get_db)
):
    """Actualizar un curso (solo admin/coordinador)"""
    curso = db.query(CursoDB).filter(CursoDB.id == curso_id).first()
    if not curso:
        raise HTTPException(404, "Curso no encontrado")

    campos_permitidos = {"nombre", "creditos", "semestre", "programa_id", "docente_id",
                         "max_estudiantes", "tipo", "horas_semanales"}
    for key, value in curso_data.items():
        if key in campos_permitidos and hasattr(curso, key):
            setattr(curso, key, value)

    db.commit()
    db.refresh(curso)
    return curso


@router.delete("/cursos/{curso_id}")
def eliminar_curso(
    curso_id: int,
    current_user: Usuario = Depends(require_role(["admin"])),
    db: Session = Depends(get_db)
):
    """Eliminar un curso (solo admin)"""
    curso = db.query(CursoDB).filter(CursoDB.id == curso_id).first()
    if not curso:
        raise HTTPException(404, "Curso no encontrado")

    db.delete(curso)
    db.commit()
    return {"mensaje": "Curso eliminado"}


# ========== SOLICITUDES DE RETIRO ==========

@router.post("/solicitar-retiro")
async def solicitar_retiro(
    matricula_id: int,
    justificacion: str,
    current_user: Usuario = Depends(require_role(["estudiante"])),
    db: Session = Depends(get_db)
):
    """Estudiante solicita retiro de un curso"""
    
    # Verificar que la matrícula existe y pertenece al estudiante
    matricula = db.query(Matricula).filter(
        Matricula.id == matricula_id,
        Matricula.estudiante_id == current_user.id
    ).first()
    
    if not matricula:
        raise HTTPException(404, "Matrícula no encontrada")
    
    # Verificar que no haya una solicitud pendiente
    solicitud_existente = db.query(SolicitudRetiro).filter(
        SolicitudRetiro.matricula_id == matricula_id,
        SolicitudRetiro.estado == "pendiente"
    ).first()
    
    if solicitud_existente:
        raise HTTPException(400, "Ya tienes una solicitud de retiro pendiente para este curso")
    
    # Crear solicitud
    solicitud = SolicitudRetiro(
        matricula_id=matricula_id,
        estudiante_id=current_user.id,
        curso_id=matricula.horario.curso_id,
        justificacion=justificacion,
        estado="pendiente",
        fecha_solicitud=datetime.utcnow()
    )
    db.add(solicitud)
    db.commit()
    
    return {
        "mensaje": "Solicitud de retiro enviada correctamente",
        "solicitud_id": solicitud.id,
        "estado": "pendiente"
    }


@router.get("/mis-solicitudes-retiro")
async def mis_solicitudes_retiro(
    current_user: Usuario = Depends(require_role(["estudiante"])),
    db: Session = Depends(get_db)
):
    """Estudiante ve sus solicitudes de retiro"""
    
    solicitudes = db.query(SolicitudRetiro).filter(
        SolicitudRetiro.estudiante_id == current_user.id
    ).order_by(SolicitudRetiro.fecha_solicitud.desc()).all()
    
    resultado = []
    for s in solicitudes:
        curso = db.query(CursoDB).filter(CursoDB.id == s.curso_id).first()
        resultado.append({
            "id": s.id,
            "curso": curso.nombre if curso else "N/A",
            "justificacion": s.justificacion,
            "estado": s.estado,
            "motivo_rechazo": s.motivo_rechazo,
            "fecha_solicitud": s.fecha_solicitud.isoformat(),
            "fecha_resolucion": s.fecha_resolucion.isoformat() if s.fecha_resolucion else None
        })
    
    return resultado


@router.get("/solicitudes-retiro")
async def listar_solicitudes_retiro(
    estado: Optional[str] = None,
    current_user: Usuario = Depends(require_role(["admin", "coordinador"])),
    db: Session = Depends(get_db)
):
    """Admin/Coordinador lista todas las solicitudes de retiro"""
    
    query = db.query(SolicitudRetiro)
    
    if estado:
        query = query.filter(SolicitudRetiro.estado == estado)
    
    solicitudes = query.order_by(SolicitudRetiro.fecha_solicitud.desc()).all()
    
    resultado = []
    for s in solicitudes:
        matricula = db.query(Matricula).filter(Matricula.id == s.matricula_id).first()
        estudiante = db.query(Usuario).filter(Usuario.id == s.estudiante_id).first()
        curso = db.query(CursoDB).filter(CursoDB.id == s.curso_id).first()
        horario = db.query(HorarioGenerado).filter(HorarioGenerado.id == matricula.horario_id).first() if matricula else None
        
        resultado.append({
            "id": s.id,
            "estudiante": f"{estudiante.nombre} {estudiante.apellido}",
            "curso": curso.nombre if curso else "N/A",
            "horario": horario.franja if horario else "N/A",
            "justificacion": s.justificacion,
            "estado": s.estado,
            "fecha_solicitud": s.fecha_solicitud.isoformat()
        })
    
    return resultado


@router.put("/solicitud-retiro/{solicitud_id}")
async def resolver_solicitud_retiro(
    solicitud_id: int,
    accion: str,  # "aprobar" o "rechazar"
    motivo_rechazo: Optional[str] = None,
    current_user: Usuario = Depends(require_role(["admin", "coordinador"])),
    db: Session = Depends(get_db)
):
    """Admin/Coordinador resuelve solicitud de retiro"""
    
    solicitud = db.query(SolicitudRetiro).filter(SolicitudRetiro.id == solicitud_id).first()
    if not solicitud:
        raise HTTPException(404, "Solicitud no encontrada")
    
    if solicitud.estado != "pendiente":
        raise HTTPException(400, f"Esta solicitud ya fue {solicitud.estado}")
    
    if accion == "aprobar":
        # Obtener matrícula y liberar cupo
        matricula = db.query(Matricula).filter(Matricula.id == solicitud.matricula_id).first()
        if matricula:
            matricula.estado = "retirado"
            
            # Liberar cupo en el horario
            horario = db.query(HorarioGenerado).filter(HorarioGenerado.id == matricula.horario_id).first()
            if horario:
                horario.cupo_disponible += 1
        
        solicitud.estado = "aprobado"
        mensaje = "Solicitud de retiro aprobada"
        
    elif accion == "rechazar":
        if not motivo_rechazo:
            raise HTTPException(400, "Debes proporcionar un motivo de rechazo")
        solicitud.estado = "rechazado"
        solicitud.motivo_rechazo = motivo_rechazo
        mensaje = "Solicitud de retiro rechazada"
    else:
        raise HTTPException(400, "Acción inválida. Use 'aprobar' o 'rechazar'")
    
    solicitud.fecha_resolucion = datetime.utcnow()
    solicitud.resuelto_por = current_user.id
    
    db.commit()
    
    return {"mensaje": mensaje, "estado": solicitud.estado}


# ========== SOLICITUDES DE CAMBIO ==========

@router.post("/solicitar-cambio")
async def solicitar_cambio(
    matricula_id: int,
    horario_deseado_id: int,
    justificacion: str,
    current_user: Usuario = Depends(require_role(["estudiante"])),
    db: Session = Depends(get_db)
):
    """Estudiante solicita cambio de horario/sección"""
    
    # Verificar matrícula
    matricula = db.query(Matricula).filter(
        Matricula.id == matricula_id,
        Matricula.estudiante_id == current_user.id
    ).first()
    
    if not matricula:
        raise HTTPException(404, "Matrícula no encontrada")
    
    # Verificar horario deseado existe y tiene cupo
    horario_deseado = db.query(HorarioGenerado).filter(HorarioGenerado.id == horario_deseado_id).first()
    if not horario_deseado:
        raise HTTPException(404, "Horario deseado no encontrado")
    
    if horario_deseado.cupo_disponible <= 0:
        raise HTTPException(400, "El horario deseado no tiene cupos disponibles")
    
    # Verificar que no sea el mismo horario
    if matricula.horario_id == horario_deseado_id:
        raise HTTPException(400, "El horario deseado es el mismo que el actual")
    
    # Verificar que no haya solicitud pendiente
    solicitud_existente = db.query(SolicitudCambio).filter(
        SolicitudCambio.matricula_id == matricula_id,
        SolicitudCambio.estado == "pendiente"
    ).first()
    
    if solicitud_existente:
        raise HTTPException(400, "Ya tienes una solicitud de cambio pendiente")
    
    # Crear solicitud
    solicitud = SolicitudCambio(
        matricula_id=matricula_id,
        estudiante_id=current_user.id,
        horario_actual_id=matricula.horario_id,
        horario_deseado_id=horario_deseado_id,
        justificacion=justificacion,
        estado="pendiente",
        fecha_solicitud=datetime.utcnow()
    )
    db.add(solicitud)
    db.commit()
    
    return {
        "mensaje": "Solicitud de cambio enviada correctamente",
        "solicitud_id": solicitud.id,
        "estado": "pendiente"
    }


@router.get("/mis-solicitudes-cambio")
async def mis_solicitudes_cambio(
    current_user: Usuario = Depends(require_role(["estudiante"])),
    db: Session = Depends(get_db)
):
    """Estudiante ve sus solicitudes de cambio"""
    
    solicitudes = db.query(SolicitudCambio).filter(
        SolicitudCambio.estudiante_id == current_user.id
    ).order_by(SolicitudCambio.fecha_solicitud.desc()).all()
    
    resultado = []
    for s in solicitudes:
        horario_actual = db.query(HorarioGenerado).filter(HorarioGenerado.id == s.horario_actual_id).first()
        horario_deseado = db.query(HorarioGenerado).filter(HorarioGenerado.id == s.horario_deseado_id).first()
        
        resultado.append({
            "id": s.id,
            "horario_actual": horario_actual.franja if horario_actual else "N/A",
            "horario_deseado": horario_deseado.franja if horario_deseado else "N/A",
            "justificacion": s.justificacion,
            "estado": s.estado,
            "motivo_rechazo": s.motivo_rechazo,
            "fecha_solicitud": s.fecha_solicitud.isoformat()
        })
    
    return resultado


@router.get("/solicitudes-cambio")
async def listar_solicitudes_cambio(
    estado: Optional[str] = None,
    current_user: Usuario = Depends(require_role(["admin", "coordinador"])),
    db: Session = Depends(get_db)
):
    """Admin/Coordinador lista todas las solicitudes de cambio"""
    
    query = db.query(SolicitudCambio)
    
    if estado:
        query = query.filter(SolicitudCambio.estado == estado)
    
    solicitudes = query.order_by(SolicitudCambio.fecha_solicitud.desc()).all()
    
    resultado = []
    for s in solicitudes:
        matricula = db.query(Matricula).filter(Matricula.id == s.matricula_id).first()
        estudiante = db.query(Usuario).filter(Usuario.id == s.estudiante_id).first()
        horario_actual = db.query(HorarioGenerado).filter(HorarioGenerado.id == s.horario_actual_id).first()
        horario_deseado = db.query(HorarioGenerado).filter(HorarioGenerado.id == s.horario_deseado_id).first()
        curso = db.query(CursoDB).filter(CursoDB.id == horario_actual.curso_id).first() if horario_actual else None
        
        resultado.append({
            "id": s.id,
            "estudiante": f"{estudiante.nombre} {estudiante.apellido}",
            "curso": curso.nombre if curso else "N/A",
            "horario_actual": horario_actual.franja if horario_actual else "N/A",
            "horario_deseado": horario_deseado.franja if horario_deseado else "N/A",
            "justificacion": s.justificacion,
            "estado": s.estado,
            "fecha_solicitud": s.fecha_solicitud.isoformat()
        })
    
    return resultado


@router.put("/solicitud-cambio/{solicitud_id}")
async def resolver_solicitud_cambio(
    solicitud_id: int,
    accion: str,  # "aprobar" o "rechazar"
    motivo_rechazo: Optional[str] = None,
    current_user: Usuario = Depends(require_role(["admin", "coordinador"])),
    db: Session = Depends(get_db)
):
    """Admin/Coordinador resuelve solicitud de cambio"""
    
    solicitud = db.query(SolicitudCambio).filter(SolicitudCambio.id == solicitud_id).first()
    if not solicitud:
        raise HTTPException(404, "Solicitud no encontrada")
    
    if solicitud.estado != "pendiente":
        raise HTTPException(400, f"Esta solicitud ya fue {solicitud.estado}")
    
    if accion == "aprobar":
        # Verificar cupo disponible en horario deseado
        horario_deseado = db.query(HorarioGenerado).filter(HorarioGenerado.id == solicitud.horario_deseado_id).first()
        if not horario_deseado or horario_deseado.cupo_disponible <= 0:
            raise HTTPException(400, "El horario deseado ya no tiene cupos disponibles")
        
        # Actualizar matrícula
        matricula = db.query(Matricula).filter(Matricula.id == solicitud.matricula_id).first()
        if matricula:
            # Liberar cupo del horario actual
            horario_actual = db.query(HorarioGenerado).filter(HorarioGenerado.id == matricula.horario_id).first()
            if horario_actual:
                horario_actual.cupo_disponible += 1
            
            # Asignar nuevo horario
            matricula.horario_id = solicitud.horario_deseado_id
            horario_deseado.cupo_disponible -= 1
        
        solicitud.estado = "aprobado"
        mensaje = "Solicitud de cambio aprobada"
        
    elif accion == "rechazar":
        if not motivo_rechazo:
            raise HTTPException(400, "Debes proporcionar un motivo de rechazo")
        solicitud.estado = "rechazado"
        solicitud.motivo_rechazo = motivo_rechazo
        mensaje = "Solicitud de cambio rechazada"
    else:
        raise HTTPException(400, "Acción inválida. Use 'aprobar' o 'rechazar'")
    
    solicitud.fecha_resolucion = datetime.utcnow()
    solicitud.resuelto_por = current_user.id
    
    db.commit()
    
    return {"mensaje": mensaje, "estado": solicitud.estado}

# ========== GENERACIÓN MASIVA ==========
@router.post("/generar-masivo")
async def generar_horarios_masivos(
    semestres: List[int] = None,  # [1, 2, 3] o None para todos
    num_versiones: int = 3,  # Número de versiones por semestre
    current_user: Usuario = Depends(require_role(["admin", "coordinador"])),
    db: Session = Depends(get_db)
):
    """
    Genera múltiples versiones de horarios por semestre
    
    - semestres: lista de semestres a generar (ej: [1, 2, 3])
    - num_versiones: número de versiones por semestre
    """
    engine = CSPEngine()
    
    # Obtener todos los cursos
    query = db.query(CursoDB)
    if semestres:
        query = query.filter(CursoDB.semestre.in_(semestres))
    cursos_db = query.all()
    
    if not cursos_db:
        raise HTTPException(404, "No hay cursos para los semestres seleccionados")
    
    # Agrupar cursos por semestre
    cursos_por_semestre = {}
    for curso in cursos_db:
        if curso.semestre not in cursos_por_semestre:
            cursos_por_semestre[curso.semestre] = []
        cursos_por_semestre[curso.semestre].append(curso)
    
    # Obtener aulas y docentes
    aulas_db = db.query(Aula).all()
    docentes_db = db.query(Docente).all()
    
    # Preparar datos para el CSP
    aulas = [{"id": a.codigo, "capacidad": a.capacidad, "tipo": a.tipo, "recursos": []} for a in aulas_db]
    docentes = [{"id": d.codigo, "nombre": f"{d.nombre} {d.apellido}", "disponibilidad": []} for d in docentes_db]
    
    resultados = {}
    todas_versiones = []
    
    for semestre, cursos_semestre in cursos_por_semestre.items():
        print(f"\n📚 Generando horarios para Semestre {semestre} - {len(cursos_semestre)} cursos")
        
        # Preparar cursos para este semestre
        cursos = []
        for c in cursos_semestre:
            cursos.append({
                "id": c.codigo,
                "nombre": c.nombre,
                "docente_id": next((d.codigo for d in docentes_db if d.id == c.docente_id), "DOC001"),
                "max_estudiantes": c.max_estudiantes,
                "tipo": c.tipo,
                "semestre": c.semestre,
                "horas_teoria": c.horas_semanales,
                "horas_laboratorio": 0,
                "horas_practica": 0,
                "minutos_por_hora": 90
            })
        
        # Generar múltiples versiones
        versiones = []
        for v in range(num_versiones):
            # Usar diferentes estrategias para cada versión
            if v == 0:
                cursos_ordenados = cursos  # Original
            elif v == 1:
                cursos_ordenados = sorted(cursos, key=lambda x: x.get('horas_semanales', 4), reverse=True)  # Más horas primero
            elif v == 2:
                cursos_ordenados = sorted(cursos, key=lambda x: x.get('tipo', 'teoria'))  # Por tipo
            else:
                import random
                random.seed(v)
                cursos_ordenados = cursos.copy()
                random.shuffle(cursos_ordenados)
            
            resultado = engine.resolver(cursos_ordenados, aulas, docentes)
            versiones.append({
                "version": v + 1,
                "asignaciones": len(resultado['horario_generado']),
                "horario": resultado['horario_generado'],
                "distribucion": resultado['estadisticas'].get('distribucion_dias', {})
            })
        
        # Guardar la mejor versión en BD (la que tiene mejor distribución)
        mejor_version = max(versiones, key=lambda x: x['asignaciones'])
        resultados[semestre] = {
            "cursos": len(cursos_semestre),
            "versiones": versiones,
            "mejor_version": mejor_version
        }
        
        # Guardar la mejor versión en la base de datos
        periodo = f"2025-{semestre}"
        
        # Limpiar horarios anteriores del periodo
        db.query(HorarioGenerado).filter(HorarioGenerado.periodo == periodo).delete()
        
        for h in mejor_version['horario']:
            curso_db = db.query(CursoDB).filter(CursoDB.codigo == h['curso_id']).first()
            docente_db = db.query(Docente).filter(Docente.codigo == h['docente_id']).first()
            aula_db = db.query(Aula).filter(Aula.codigo == h['aula_id']).first()
            
            if curso_db:
                hg = HorarioGenerado(
                    periodo=periodo,
                    curso_id=curso_db.id,
                    docente_id=docente_db.id if docente_db else None,
                    aula_id=aula_db.id if aula_db else None,
                    franja=h['franja'],
                    seccion=chr(65 + (semestre - 1)),  # A, B, C según semestre
                    cupo_maximo=curso_db.max_estudiantes,
                    cupo_disponible=curso_db.max_estudiantes
                )
                db.add(hg)
        
        db.commit()
        print(f"✅ Semestre {semestre}: guardada mejor versión con {mejor_version['asignaciones']} asignaciones")
        todas_versiones.append(mejor_version)
    
    return {
        "success": True,
        "semestres_generados": len(resultados),
        "total_horarios": sum(len(db.query(HorarioGenerado).filter(HorarioGenerado.periodo == f"2025-{s}").all()) for s in resultados.keys()),
        "detalle": resultados
    }


@router.post("/generar", response_model=HorarioResponse)
async def generar_horario(request: HorarioRequest, db: Session = Depends(get_db)):
    """Genera horarios usando el motor CSP"""
    logger.info("="*50)
    logger.info("INICIANDO GENERACIÓN DE HORARIOS")
    
    engine = CSPEngine()
    
    # Convertir a diccionarios de forma segura
    cursos = []
    for c in request.cursos:
        if hasattr(c, 'model_dump'):
            cursos.append(c.model_dump())
        else:
            cursos.append(c)
    
    aulas = []
    for a in request.aulas:
        if hasattr(a, 'model_dump'):
            aulas.append(a.model_dump())
        else:
            aulas.append(a)
    
    docentes = []
    for d in request.docentes:
        if hasattr(d, 'model_dump'):
            docentes.append(d.model_dump())
        else:
            docentes.append(d)
    
    logger.info(f"Ejecutando CSP con {len(cursos)} cursos...")
    
    resultado = engine.resolver(cursos, aulas, docentes)
    
    # ========== GUARDAR EN BASE DE DATOS ==========
    if resultado['horario_generado']:
        periodo = getattr(request, 'periodo', "2025-1")
        
        try:
            # 1. Primero buscar los horarios del periodo que se van a eliminar
            from models.db_models import Matricula, SolicitudRetiro, SolicitudCambio, HistorialCambios
            horarios_a_eliminar = db.query(HorarioGenerado).filter(HorarioGenerado.periodo == periodo).all()
            ids_horarios = [h.id for h in horarios_a_eliminar]
            
            if ids_horarios:
                # 2. Eliminar historial de cambios asociado a estos horarios
                db.query(HistorialCambios).filter(HistorialCambios.horario_id.in_(ids_horarios)).delete(synchronize_session=False)
                
                # 3. Eliminar solicitudes de cambio asociadas a estos horarios
                db.query(SolicitudCambio).filter(
                    (SolicitudCambio.horario_actual_id.in_(ids_horarios)) |
                    (SolicitudCambio.horario_deseado_id.in_(ids_horarios))
                ).delete(synchronize_session=False)

                # 4. Eliminar solicitudes de retiro y matrículas
                matriculas = db.query(Matricula).filter(Matricula.horario_id.in_(ids_horarios)).all()
                ids_matriculas = [m.id for m in matriculas]
                
                if ids_matriculas:
                    db.query(SolicitudRetiro).filter(SolicitudRetiro.matricula_id.in_(ids_matriculas)).delete(synchronize_session=False)
                    db.query(SolicitudCambio).filter(SolicitudCambio.matricula_id.in_(ids_matriculas)).delete(synchronize_session=False)
                    db.query(Matricula).filter(Matricula.id.in_(ids_matriculas)).delete(synchronize_session=False)
                
                db.commit()
                logger.info(f"✅ Eliminadas {len(ids_matriculas)} matrículas y dependencias")
                
                # 5. Finalmente, eliminar los horarios
                horarios_eliminados = db.query(HorarioGenerado).filter(HorarioGenerado.periodo == periodo).delete(synchronize_session=False)
                db.commit()
                logger.info(f"✅ Eliminados {horarios_eliminados} horarios anteriores")
            
            # 3. Insertar nuevos horarios
            for h in resultado['horario_generado']:
                curso_db = db.query(CursoDB).filter(CursoDB.codigo == h['curso_id']).first()
                docente_db = db.query(Docente).filter(Docente.codigo == h['docente_id']).first()
                aula_db = db.query(Aula).filter(Aula.codigo == h['aula_id']).first()
                
                if not curso_db:
                    logger.warning(f"Curso no encontrado: {h['curso_id']}")
                    continue
                
                hg = HorarioGenerado(
                    periodo=periodo,
                    curso_id=curso_db.id,
                    docente_id=docente_db.id if docente_db else None,
                    aula_id=aula_db.id if aula_db else None,
                    franja=h['franja'],
                    seccion=h.get('seccion', 'A'),
                    cupo_maximo=curso_db.max_estudiantes,
                    cupo_disponible=curso_db.max_estudiantes
                )
                db.add(hg)
            
            db.commit()
            logger.info(f"✅ Guardados {len(resultado['horario_generado'])} horarios en BD")
            
        except Exception as e:
            logger.error(f"Error guardando en BD: {e}")
            db.rollback()
    else:
        logger.warning("No hay horarios para guardar")
    
    logger.info("="*50)
    return resultado

# ========== GUARDAR SOLUCIÓN CSP EN BD ==========
class GuardarSolucionRequest(BaseModel):
    semestre: int = Field(..., ge=1, le=12)
    horario_generado: List[Dict[str, Any]] = Field(..., min_items=1)

@router.post("/guardar-solucion-horario")
def guardar_solucion_horario(
    request: GuardarSolucionRequest,  # ← Usa 'request', no 'data'
    current_user: Usuario = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    Persiste en BD los horarios de una solución CSP.
    Devuelve lista de IDs reales para llamar a /matricular.
    """
    periodo = f"2025-{request.semestre}"  # ← CORREGIDO: request.semestre
    logger.info(
        f"[guardar-solucion] usuario={current_user.email} "
        f"semestre={request.semestre} periodo={periodo} items={len(request.horario_generado)}"
    )
    
    try:
        # 1. Cancelar matrículas previas del estudiante en este periodo
        matriculas_previas = db.query(Matricula).filter(
            Matricula.estudiante_id == current_user.id,
            Matricula.periodo == periodo,
            Matricula.estado == "activa"
        ).all()
        for m in matriculas_previas:
            if m.horario:
                m.horario.cupo_disponible += 1
            db.delete(m)
        db.commit()
        logger.info(f"[guardar-solucion] canceladas {len(matriculas_previas)} matrículas previas")

        # 2. Procesar cada asignación del CSP
        ids_creados = []
        no_encontrados = []

        for h in request.horario_generado:  # ← CORREGIDO: request.horario_generado
            # Resolver CURSO
            curso_db = None
            curso_key = h.get("curso_id") or h.get("codigo")
            if curso_key:
                curso_db = db.query(CursoDB).filter(
                    CursoDB.codigo == curso_key.strip().upper()
                ).first()
            
            if not curso_db:
                logger.warning(f"[guardar-solucion] curso no encontrado: '{curso_key}'")
                no_encontrados.append(str(curso_key))
                continue

            # Resolver DOCENTE
            docente_db = None
            doc_key = h.get("docente_id")
            if doc_key:
                try:
                    docente_db = db.query(Docente).filter(Docente.id == int(doc_key)).first()
                except (ValueError, TypeError):
                    docente_db = db.query(Docente).filter(Docente.codigo == str(doc_key)).first()
            
            if not docente_db and curso_db.docente_id:
                docente_db = db.query(Docente).filter(Docente.id == curso_db.docente_id).first()

            # Resolver AULA
            aula_db = None
            aula_key = h.get("aula_id")
            if aula_key:
                if isinstance(aula_key, str):
                    aula_db = db.query(Aula).filter(Aula.codigo == aula_key.strip()).first()
                else:
                    aula_db = db.query(Aula).filter(Aula.id == aula_key).first()

            # Construir franja
            franja = h.get("franja")
            if not franja:
                dia = h.get("dia", "Lunes")
                fi = h.get("franja_inicio", "07:00")
                ff = h.get("franja_fin", "08:30")
                franja = f"{dia} {fi} - {ff}"

            # Crear nuevo horario
            hg = HorarioGenerado(
                periodo=periodo,
                curso_id=curso_db.id,
                docente_id=docente_db.id if docente_db else None,
                aula_id=aula_db.id if aula_db else None,
                franja=franja,
                seccion=h.get("seccion", "A"),
                cupo_maximo=curso_db.max_estudiantes,
                cupo_disponible=curso_db.max_estudiantes
            )
            db.add(hg)
            db.flush()
            ids_creados.append(hg.id)
            logger.info(f"[guardar-solucion] creado id={hg.id} curso={curso_db.codigo}")

        db.commit()
        logger.info(f"[guardar-solucion] ✅ {len(ids_creados)} ids creados")

        if not ids_creados:
            raise HTTPException(
                status_code=422,
                detail=f"No se encontraron cursos: {no_encontrados}"
            )

        return {
            "ok": True,
            "periodo": periodo,
            "horario_ids": ids_creados,
            "total": len(ids_creados),
            "advertencias": no_encontrados
        }

    except HTTPException:
        raise
    except Exception as exc:
        db.rollback()
        logger.error(f"[guardar-solucion] ❌ {exc}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error al guardar solución: {str(exc)}")
def guardar_solucion_horario(
    request: GuardarSolucionRequest,  # ← Se llama 'request', no 'data'
    current_user: Usuario = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    Persiste en BD los horarios de una solución CSP (CSPEngineProfesional).
    Devuelve lista de IDs reales para llamar a /matricular.
    """
    periodo = f"2025-{request.semestre}"  # ← Cambiar data a request
    logger.info(
        f"[guardar-solucion] usuario={current_user.email} "
        f"semestre={request.semestre} periodo={periodo} items={len(request.horario_generado)}"
    )
    if request.horario_generado:  # ← Cambiar data a request
        logger.info(f"[guardar-solucion] sample={request.horario_generado[0]}")

    try:
        # 1. Cancelar matrículas previas del estudiante en este periodo
        matriculas_previas = db.query(Matricula).filter(
            Matricula.estudiante_id == current_user.id,
            Matricula.periodo == periodo,
            Matricula.estado == "activa"
        ).all()
        for m in matriculas_previas:
            if m.horario:
                m.horario.cupo_disponible += 1
            db.delete(m)
        db.flush()
        logger.info(f"[guardar-solucion] canceladas {len(matriculas_previas)} matrículas previas")

        # 2. Procesar cada asignación del CSP
        ids_creados = []
        no_encontrados = []

        for h in request.horario_generado:  # ← Cambiar data a request
            # ─ Resolver CURSO ─
            curso_db = None
            curso_key = h.get("curso_id") or h.get("codigo")
            if isinstance(curso_key, str):
                curso_db = db.query(CursoDB).filter(
                    CursoDB.codigo == curso_key.strip().upper()
                ).first()
            if not curso_db:
                try:
                    curso_db = db.query(CursoDB).filter(
                        CursoDB.id == int(curso_key)
                    ).first()
                except (ValueError, TypeError):
                    pass
            if not curso_db:
                logger.warning(f"[guardar-solucion] curso no encontrado: '{curso_key}'")
                no_encontrados.append(str(curso_key))
                continue

            # ─ Resolver DOCENTE ─
            docente_db = None
            doc_key = h.get("docente_id")
            if doc_key is not None:
                try:
                    docente_db = db.query(Docente).filter(
                        Docente.id == int(doc_key)
                    ).first()
                except (ValueError, TypeError):
                    docente_db = db.query(Docente).filter(
                        Docente.codigo == str(doc_key)
                    ).first()
            if not docente_db and curso_db.docente_id:
                docente_db = db.query(Docente).filter(
                    Docente.id == curso_db.docente_id
                ).first()

            # ─ Resolver AULA ─
            aula_db = None
            aula_key = h.get("aula_id")
            if isinstance(aula_key, str):
                aula_db = db.query(Aula).filter(
                    Aula.codigo == aula_key.strip()
                ).first()
            if not aula_db:
                try:
                    aula_db = db.query(Aula).filter(
                        Aula.id == int(aula_key)
                    ).first()
                except (ValueError, TypeError):
                    pass

            # ─ Construir franja ─
            franja = h.get("franja")
            if not franja:
                dia = h.get("dia", "Lunes")
                fi = h.get("franja_inicio", "07:00")
                ff = h.get("franja_fin", "08:30")
                franja = f"{dia} {fi} - {ff}"

            # ─ Reusar si ya existe el horario ─
            existente = db.query(HorarioGenerado).filter(
                HorarioGenerado.periodo == periodo,
                HorarioGenerado.curso_id == curso_db.id,
                HorarioGenerado.franja == franja,
            ).first()
            if existente:
                logger.info(f"[guardar-solucion] reutilizando id={existente.id} ({curso_db.codigo})")
                ids_creados.append(existente.id)
                continue

            # ─ Insertar nuevo ─
            hg = HorarioGenerado(
                periodo=periodo,
                curso_id=curso_db.id,
                docente_id=docente_db.id if docente_db else None,
                aula_id=aula_db.id if aula_db else None,
                franja=franja,
                seccion=h.get("seccion", "A"),
                cupo_maximo=curso_db.max_estudiantes,
                cupo_disponible=curso_db.max_estudiantes,
            )
            db.add(hg)
            db.flush()
            ids_creados.append(hg.id)
            logger.info(f"[guardar-solucion] creado id={hg.id} curso={curso_db.codigo} franja='{franja}'")

        db.commit()
        logger.info(f"[guardar-solucion] ✅ {len(ids_creados)} ids | no encontrados={no_encontrados}")

        if not ids_creados:
            raise HTTPException(
                status_code=422,
                detail=(
                    f"No se encontraron los cursos del CSP en la base de datos. "
                    f"Cursos buscados: {no_encontrados}. "
                    f"Asegúrate de generar el horario con cursos del semestre {request.semestre}."
                )
            )

        return {
            "ok": True,
            "periodo": periodo,
            "horario_ids": ids_creados,
            "total": len(ids_creados),
            "advertencias": no_encontrados,
        }

    except HTTPException:
        raise
    except Exception as exc:
        db.rollback()
        logger.error(f"[guardar-solucion] ❌ {exc}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error al guardar solución: {str(exc)}")

# ========== HORARIOS DISPONIBLES ==========
@router.get("/horarios-disponibles")
def horarios_disponibles(
    periodo: str = "2025-1", 
    programa_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """Retorna horarios disponibles para matrícula"""
    query = db.query(HorarioGenerado).filter(
        HorarioGenerado.periodo == periodo,
        HorarioGenerado.cupo_disponible > 0
    )
    
    horarios = query.all()
    
    resultado = []
    for h in horarios:
        curso = db.query(CursoDB).filter(CursoDB.id == h.curso_id).first() if h.curso_id else None
        
        # Filtrar por programa si se especifica
        if programa_id and curso and curso.programa_id != programa_id:
            continue
        
        resultado.append({
            "id": h.id,
            "curso_id": h.curso_id,
            "curso_codigo": curso.codigo if curso else "N/A",
            "curso_nombre": curso.nombre if curso else "Desconocido",
            "creditos": curso.creditos if curso else 0,
            "docente_id": h.docente_id,
            "docente_nombre": get_docente_nombre(db, h.docente_id) if h.docente_id else "N/A",
            "aula_id": h.aula_id,
            "franja": h.franja,
            "seccion": h.seccion,
            "cupo_maximo": h.cupo_maximo,
            "cupo_disponible": h.cupo_disponible,
            "ocupacion": round((1 - h.cupo_disponible/h.cupo_maximo) * 100) if h.cupo_maximo > 0 else 0
        })
    
    return resultado

# ========== MATRÍCULA ==========
@router.post("/matricular")
def matricular(
    data: MatriculaInscripcion,
    current_user: Usuario = Depends(require_role(["estudiante"])),
    db: Session = Depends(get_db)
):
    """
    Estudiante se matricula en cursos. Validaciones:
    - Cupos disponibles
    - Sin cruces de horario
    - Sin duplicados
    - Límite de cursos por semestre
    """
    periodo_actual = "2025-1"
    max_cursos = 10  # Aumentado para permitir carga completa de semestre
    
    # Obtener horarios seleccionados
    horarios = db.query(HorarioGenerado).filter(
        HorarioGenerado.id.in_(data.horario_ids)
    ).all()
    
    if len(horarios) != len(data.horario_ids):
        raise HTTPException(400, "Algunos horarios no existen")
    
    # Validar cupos disponibles
    for h in horarios:
        if h.cupo_disponible <= 0:
            curso = db.query(CursoDB).filter(CursoDB.id == h.curso_id).first()
            raise HTTPException(400, f"Curso {curso.codigo if curso else h.curso_id} sin cupos disponibles")
    
    # Validar no duplicados (mismo curso)
    cursos_ids = [h.curso_id for h in horarios]
    if len(cursos_ids) != len(set(cursos_ids)):
        raise HTTPException(400, "No puedes inscribir el mismo curso dos veces")
    
    # Validar cruce de horarios
    franjas = [h.franja for h in horarios]
    if len(franjas) != len(set(franjas)):
        raise HTTPException(400, "Cruce de horario detectado. Selecciona franjas diferentes.")
    
    # Verificar matrículas previas del estudiante
    matriculas_previas = db.query(Matricula).filter(
        Matricula.estudiante_id == current_user.id,
        Matricula.periodo == periodo_actual,
        Matricula.estado == "activa"
    ).all()
    
    # Verificar límite de cursos
    if len(matriculas_previas) + len(horarios) > max_cursos:
        raise HTTPException(400, f"Máximo {max_cursos} cursos por semestre")
    
    # Verificar cursos ya matriculados
    cursos_previos = {m.horario.curso_id for m in matriculas_previas if m.horario}
    cursos_nuevos = set(cursos_ids)
    conflictos = cursos_previos & cursos_nuevos
    
    if conflictos:
        nombres_conflictos = []
        for curso_id in conflictos:
            curso = db.query(CursoDB).filter(CursoDB.id == curso_id).first()
            nombres_conflictos.append(curso.codigo if curso else str(curso_id))
        raise HTTPException(400, f"Ya estás matriculado en: {', '.join(nombres_conflictos)}")
    
    # Crear matrículas y descontar cupos
    nuevas_matriculas = []
    for h in horarios:
        h.cupo_disponible -= 1
        
        matricula = Matricula(
            estudiante_id=current_user.id,
            horario_id=h.id,
            periodo=periodo_actual,
            estado="activa"
        )
        db.add(matricula)
        nuevas_matriculas.append(matricula)
    
    db.commit()
    
    logger.info(f"Estudiante {current_user.email} matriculado en {len(nuevas_matriculas)} cursos")
    
    return {
        "mensaje": "Matrícula exitosa",
        "matriculas_realizadas": len(nuevas_matriculas),
        "periodo": periodo_actual,
        "cursos": [
            {
                "curso_id": h.curso_id,
                "curso_nombre": get_curso_nombre(db, h.curso_id),
                "franja": h.franja,
                "aula": h.aula_id,
                "seccion": h.seccion
            }
            for h in horarios
        ]
    }

@router.get("/mi-horario")
def mi_horario(
    current_user: Usuario = Depends(require_role(["estudiante", "admin", "coordinador", "docente"])),
    db: Session = Depends(get_db)
):
    """Retorna el horario personal del estudiante autenticado"""
    periodo_actual = "2025-1"
    
    print(f"🔍 Buscando matrículas para estudiante ID: {current_user.id}")
    
    # Obtener matrículas activas
    matriculas = db.query(Matricula).filter(
        Matricula.estudiante_id == current_user.id,
        Matricula.periodo == periodo_actual,
        Matricula.estado == "activa"
    ).all()
    
    print(f"📋 Matrículas encontradas: {len(matriculas)}")
    
    horario_personal = []
    for matricula in matriculas:
        print(f"   - Matrícula ID: {matricula.id}, Horario ID: {matricula.horario_id}")
        
        # Obtener el horario asociado
        h = db.query(HorarioGenerado).filter(HorarioGenerado.id == matricula.horario_id).first()
        
        if not h:
            print(f"   ⚠️ Horario no encontrado para ID: {matricula.horario_id}")
            continue
        
        print(f"   ✅ Horario encontrado: {h.franja}")
        
        # Obtener el curso
        curso = db.query(CursoDB).filter(CursoDB.id == h.curso_id).first()
        
        horario_personal.append({
            "matricula_id": matricula.id,
            "curso_id": curso.codigo if curso else str(h.curso_id),
            "curso_codigo": curso.codigo if curso else "N/A",
            "curso_nombre": curso.nombre if curso else "Desconocido",
            "creditos": curso.creditos if curso else 0,
            "docente": get_docente_nombre(db, h.docente_id),
            "aula": h.aula_id,
            "franja": h.franja,
            "seccion": h.seccion,
            "fecha_matricula": matricula.created_at.isoformat() if matricula.created_at else None
        })
    
    print(f"✅ Total cursos a mostrar: {len(horario_personal)}")
    
    return {
        "estudiante": f"{current_user.nombre} {current_user.apellido}",
        "email": current_user.email,
        "periodo": periodo_actual,
        "total_cursos": len(horario_personal),
        "total_creditos": sum(h['creditos'] for h in horario_personal),
        "horario": horario_personal
    }

@router.get("/mis-matriculas")
def mis_matriculas(
    current_user: Usuario = Depends(require_role(["estudiante"])),
    db: Session = Depends(get_db)
):
    """Retorna historial de matrículas del estudiante"""
    matriculas = db.query(Matricula).filter(
        Matricula.estudiante_id == current_user.id
    ).order_by(Matricula.created_at.desc()).all()
    
    resultado = []
    for m in matriculas:
        curso_nombre = "N/A"
        if m.horario and m.horario.curso_id:
            curso = db.query(CursoDB).filter(CursoDB.id == m.horario.curso_id).first()
            curso_nombre = curso.nombre if curso else "N/A"
        
        resultado.append({
            "id": m.id,
            "periodo": m.periodo,
            "curso": curso_nombre,
            "estado": m.estado,
            "nota_final": m.nota_final,
            "fecha_matricula": m.created_at.isoformat() if m.created_at else None
        })
    
    return resultado

@router.delete("/mis-matriculas/{matricula_id}")
def cancelar_matricula(
    matricula_id: int,
    current_user: Usuario = Depends(require_role(["estudiante"])),
    db: Session = Depends(get_db)
):
    """Cancelar una matrícula (libera cupo)"""
    matricula = db.query(Matricula).filter(
        Matricula.id == matricula_id,
        Matricula.estudiante_id == current_user.id,
        Matricula.estado == "activa"
    ).first()
    
    if not matricula:
        raise HTTPException(404, "Matrícula no encontrada")
    
    # Liberar cupo
    if matricula.horario:
        matricula.horario.cupo_disponible += 1
    
    matricula.estado = "cancelada"
    db.commit()
    
    logger.info(f"Matrícula {matricula_id} cancelada por estudiante {current_user.email}")
    
    return {"mensaje": "Matrícula cancelada exitosamente"}

# ========== REPORTES ==========
@router.get("/reportes/ocupacion")
def reporte_ocupacion(db: Session = Depends(get_db)):
    """Dashboard: ocupación por aula y franja"""
    horarios = db.query(HorarioGenerado).filter(HorarioGenerado.cupo_disponible >= 0).all()
    
    ocupacion_por_franja = {}
    ocupacion_por_aula = {}
    cursos_por_docente = {}
    
    for h in horarios:
        # Por franja
        franja = h.franja
        ocupacion_por_franja[franja] = ocupacion_por_franja.get(franja, 0) + 1
        
        # Por aula
        if h.aula_id:
            aula_nombre = f"Aula {h.aula_id}"
            ocupacion_por_aula[aula_nombre] = ocupacion_por_aula.get(aula_nombre, 0) + 1
        
        # Por docente
        if h.docente_id:
            docente_nombre = get_docente_nombre(db, h.docente_id)
            cursos_por_docente[docente_nombre] = cursos_por_docente.get(docente_nombre, 0) + 1
    
    # Calcular ocupación total
    total_cupos = sum(h.cupo_maximo for h in horarios)
    total_ocupados = sum(h.cupo_maximo - h.cupo_disponible for h in horarios)
    porcentaje_ocupacion = (total_ocupados / total_cupos * 100) if total_cupos > 0 else 0
    
    return {
        "total_franjas_asignadas": len(horarios),
        "total_cupos_disponibles": sum(h.cupo_disponible for h in horarios),
        "porcentaje_ocupacion_general": round(porcentaje_ocupacion, 2),
        "ocupacion_por_franja": ocupacion_por_franja,
        "aulas_mas_usadas": sorted(ocupacion_por_aula.items(), key=lambda x: x[1], reverse=True)[:5],
        "docentes_con_mas_cursos": sorted(cursos_por_docente.items(), key=lambda x: x[1], reverse=True)[:5]
    }

@router.get("/reportes/cursos-no-asignados")
def cursos_no_asignados(periodo: str = "2025-1", db: Session = Depends(get_db)):
    """Reporte de cursos que no tienen horario asignado"""
    cursos_con_horario = db.query(HorarioGenerado.curso_id).filter(
        HorarioGenerado.periodo == periodo
    ).distinct().all()
    
    cursos_con_horario_ids = [c[0] for c in cursos_con_horario if c[0]]
    
    cursos_sin_asignar = db.query(CursoDB).filter(
        ~CursoDB.id.in_(cursos_con_horario_ids) if cursos_con_horario_ids else True
    ).all()
    
    return {
        "periodo": periodo,
        "total_cursos_sin_asignar": len(cursos_sin_asignar),
        "cursos": [
            {
                "codigo": c.codigo,
                "nombre": c.nombre,
                "semestre": c.semestre,
                "programa_id": c.programa_id
            }
            for c in cursos_sin_asignar
        ]
    }

# =========================
# EXPORTACIÓN A PDF
# =========================

@router.get("/exportar/horario-pdf")
async def exportar_horario_pdf(
    periodo: str = "2025-1",
    programa_id: int = None,
    current_user: Usuario = Depends(require_role(["admin", "coordinador", "estudiante"])),
    db: Session = Depends(get_db)
):
    """Exporta el horario a PDF"""
    
    # Obtener horarios
    query = db.query(HorarioGenerado).filter(HorarioGenerado.periodo == periodo)
    
    horarios = query.all()
    
    if not horarios:
        raise HTTPException(404, "No hay horarios para el período seleccionado")
    
    # Crear PDF en memoria
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), title=f"Horario {periodo}")
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1e3a8a'),
        alignment=1,  # Centrado
        spaceAfter=20
    )
    
    # Título
    elementos = []
    titulo = Paragraph(f"<b>HORARIO ACADÉMICO {periodo}</b>", title_style)
    elementos.append(titulo)
    elementos.append(Spacer(1, 0.5 * cm))
    
    # Datos de la tabla
    dias = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']
    horas = ['07:30-09:00', '09:10-10:40', '10:50-12:20', '12:30-14:00', '14:10-15:40', '15:50-17:20']
    
    # Crear matriz de datos
    data = [['Hora / Día'] + dias]
    
    for hora in horas:
        fila = [hora]
        for dia in dias:
            horario = next(
                (h for h in horarios if f"{dia} {hora}" in h.franja),
                None
            )
            if horario:
                curso = db.query(CursoDB).filter(CursoDB.id == horario.curso_id).first()
                fila.append(f"{curso.codigo}\n{curso.nombre[:20]}")
            else:
                fila.append("-")
        data.append(fila)
    
    # Crear tabla
    tabla = Table(data, colWidths=[1.5*cm] + [2.8*cm]*5, repeatRows=1)
    
    # Estilo de la tabla
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    
    elementos.append(tabla)
    
    # Generar PDF
    doc.build(elementos)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=horario_{periodo}.pdf"}
    )


# =========================
# EXPORTACIÓN A EXCEL
# =========================

@router.get("/exportar/horario-excel")
async def exportar_horario_excel(
    periodo: str = "2025-1",
    current_user: Usuario = Depends(require_role(["admin", "coordinador", "estudiante"])),
    db: Session = Depends(get_db)
):
    """Exporta el horario a Excel"""
    
    # Obtener horarios
    horarios = db.query(HorarioGenerado).filter(HorarioGenerado.periodo == periodo).all()
    
    if not horarios:
        raise HTTPException(404, "No hay horarios para el período seleccionado")
    
    # Crear libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Horario {periodo}"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    dias = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']
    ws['A1'] = 'Hora / Día'
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = header_alignment
    
    for i, dia in enumerate(dias, start=2):
        cell = ws.cell(row=1, column=i)
        cell.value = dia
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Horas
    horas = ['07:30-09:00', '09:10-10:40', '10:50-12:20', '12:30-14:00', '14:10-15:40', '15:50-17:20']
    
    for i, hora in enumerate(horas, start=2):
        ws.cell(row=i, column=1, value=hora)
        ws.cell(row=i, column=1).alignment = Alignment(horizontal="center")
    
    # Llenar datos
    for i, hora in enumerate(horas, start=2):
        for j, dia in enumerate(dias, start=2):
            horario = next(
                (h for h in horarios if f"{dia} {hora}" in h.franja),
                None
            )
            if horario:
                curso = db.query(CursoDB).filter(CursoDB.id == horario.curso_id).first()
                ws.cell(row=i, column=j, value=f"{curso.codigo}\n{curso.nombre}")
                ws.cell(row=i, column=j).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                ws.cell(row=i, column=j, value="-")
                ws.cell(row=i, column=j).alignment = Alignment(horizontal="center")
    
    # Ajustar anchos de columna
    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 25
    ws.row_dimensions[1].height = 30
    
    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=horario_{periodo}.xlsx"}
    )


# =========================
# REPORTE DE ESTUDIANTE (PDF)
# =========================

@router.get("/exportar/mi-horario-pdf")
async def exportar_mi_horario_pdf(
    current_user: Usuario = Depends(require_role(["estudiante"])),
    db: Session = Depends(get_db)
):
    """Exporta el horario personal del estudiante a PDF"""
    
    # Obtener matrículas del estudiante
    matriculas = db.query(Matricula).filter(
        Matricula.estudiante_id == current_user.id,
        Matricula.periodo == "2025-1",
        Matricula.estado == "activa"
    ).all()
    
    if not matriculas:
        raise HTTPException(404, "No tienes cursos matriculados")
    
    # Crear PDF personalizado
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, title=f"Horario {current_user.nombre}")
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=14,
        textColor=colors.HexColor('#1e3a8a'),
        alignment=1
    )
    
    elementos = []
    
    # Título con datos del estudiante
    titulo = Paragraph(f"<b>HORARIO ACADÉMICO</b><br/><font size='10'>{current_user.nombre} {current_user.apellido}<br/>{current_user.email}</font>", title_style)
    elementos.append(titulo)
    elementos.append(Spacer(1, 0.5 * cm))
    
    # Lista de cursos
    data = [['Código', 'Curso', 'Horario', 'Aula']]
    
    for m in matriculas:
        if m.horario:
            curso = db.query(CursoDB).filter(CursoDB.id == m.horario.curso_id).first()
            data.append([
                curso.codigo if curso else 'N/A',
                curso.nombre if curso else 'N/A',
                m.horario.franja,
                m.horario.aula_id or 'N/A'
            ])
    
    tabla = Table(data, colWidths=[2.5*cm, 7*cm, 4*cm, 2.5*cm])
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elementos.append(tabla)
    
    doc.build(elementos)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=mi_horario.pdf"}
    )

# ========== AUTENTICACIÓN ==========
@router.post("/register")
def register_user(data: UsuarioCreate, db: Session = Depends(get_db)):
    """Registro de nuevo usuario"""
    existing = db.query(Usuario).filter(Usuario.email == data.email).first()
    if existing:
        raise HTTPException(400, "El email ya está registrado")
    
    hashed_password = get_password_hash(data.password)
    
    user = Usuario(
        email=data.email,
        hashed_password=hashed_password,
        nombre=data.nombre,
        apellido=data.apellido,
        rol=data.rol,
        activo=True
    )
    
    db.add(user)
    db.commit()
    db.refresh(user)
    
    logger.info(f"Nuevo usuario registrado: {user.email} - Rol: {user.rol}")
    
    return {
        "mensaje": "Usuario creado exitosamente",
        "usuario": {
            "id": user.id,
            "email": user.email,
            "nombre": user.nombre,
            "apellido": user.apellido,
            "rol": user.rol
        }
    }

@router.post("/login")
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    """
    Login con OAuth2. Recibe username (email) y password.
    Retorna token JWT para usar en header: Authorization: Bearer <token>
    """
    user = authenticate_user(db, form_data.username, form_data.password)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Email o contraseña incorrectos",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user.email, "rol": user.rol}, 
        expires_delta=access_token_expires
    )
    
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "usuario": {
            "id": user.id,
            "email": user.email,
            "nombre": user.nombre,
            "apellido": user.apellido,
            "rol": user.rol
        }
    }

# =========================
# RECUPERACIÓN DE CONTRASEÑA
# =========================

@router.post("/forgot-password")
async def forgot_password(
    request: ForgotPasswordRequest,
    db: Session = Depends(get_db)
):
    """
    Solicita recuperación de contraseña.
    Envía email con enlace para restablecer contraseña.
    """
    # Verificar si el usuario existe
    user = db.query(Usuario).filter(Usuario.email == request.email).first()
    if not user:
        # Por seguridad, no revelamos si el email existe o no
        return {
            "message": "Si el email está registrado, recibirás un enlace de recuperación",
            "success": True
        }
    
    # Generar token
    token = generar_token_recuperacion(request.email)
    
    # En desarrollo, devolvemos el token para pruebas
    # En producción, enviar email
    reset_link = f"http://localhost:5173/reset-password?token={token}"
    
    print(f"\n🔐 ENLACE DE RECUPERACIÓN (modo desarrollo):")
    print(f"   Email: {request.email}")
    print(f"   Token: {token}")
    print(f"   Link: {reset_link}\n")
    
    # Escribir a un log local para simular el envío de email sin configuración SMTP
    import os
    with open(os.path.join(os.path.dirname(__file__), "emails.log"), "a", encoding="utf-8") as f:
        f.write(f"[{datetime.now().isoformat()}] EMAIL PARA: {request.email}\n")
        f.write(f"Asunto: Recuperación de Contraseña OptiAcademic\n")
        f.write(f"Mensaje: Haz clic en el siguiente enlace para restablecer tu contraseña:\n{reset_link}\n")
        f.write("-" * 50 + "\n")
    
    return {
        "message": "Si el email está registrado, recibirás un enlace de recuperación",
        "success": True,
        "dev_token": token  # Solo para desarrollo
    }


@router.post("/reset-password")
async def reset_password(
    request: ResetPasswordRequest,
    db: Session = Depends(get_db)
):
    """
    Restablece la contraseña usando el token de recuperación.
    """
    # Verificar token
    email = verificar_token_recuperacion(request.token)
    if not email:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Token inválido o expirado. Solicita un nuevo enlace de recuperación."
        )
    
    # Validar nueva contraseña
    if len(request.new_password) < 6:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="La contraseña debe tener al menos 6 caracteres"
        )
    
    # Actualizar contraseña
    user = db.query(Usuario).filter(Usuario.email == email).first()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Usuario no encontrado"
        )
    
    user.hashed_password = get_password_hash(request.new_password)
    db.commit()
    
    # Eliminar token usado
    if request.token in reset_tokens:
        del reset_tokens[request.token]
    
    return {
        "message": "Contraseña restablecida exitosamente",
        "success": True
    }


@router.get("/verify-reset-token/{token}")
async def verify_reset_token(token: str):
    """
    Verifica si un token de recuperación es válido.
    """
    email = verificar_token_recuperacion(token)
    if not email:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Token inválido o expirado"
        )
    return {"valid": True, "email": email}


@router.get("/me")
def me(current_user: Usuario = Depends(get_current_active_user)):
    """Retorna información del usuario autenticado"""
    return {
        "id": current_user.id,
        "email": current_user.email,
        "nombre": current_user.nombre,
        "apellido": current_user.apellido,
        "rol": current_user.rol,
        "activo": current_user.activo
    }

# ========== RUTAS PROTEGIDAS POR ROL ==========
@router.get("/admin/dashboard")
def admin_dashboard(
    current_user: Usuario = Depends(require_role(["admin"])),
    db: Session = Depends(get_db)
):
    """Solo accesible por administradores"""
    from sqlalchemy import func
    total_usuarios   = db.query(func.count(Usuario.id)).scalar()
    total_cursos     = db.query(func.count(CursoDB.id)).scalar()
    total_matriculas = db.query(func.count(Matricula.id)).scalar()
    return {
        "mensaje": f"Bienvenido admin {current_user.nombre}",
        "stats": {
            "total_usuarios":   total_usuarios,
            "total_cursos":     total_cursos,
            "total_matriculas": total_matriculas
        }
    }

@router.get("/coordinador/horarios")
def coord_horarios(current_user: Usuario = Depends(require_role(["admin", "coordinador"]))):
    """Accesible por admin y coordinador"""
    return {
        "mensaje": f"Bienvenido coordinador {current_user.nombre}",
        "acciones": ["generar_horario", "editar_horario", "asignar_docentes"]
    }

@router.get("/docente/mis-cursos")
def docente_mis_cursos(
    current_user: Usuario = Depends(require_role(["docente", "admin"])),
    db: Session = Depends(get_db)
):
    """Docente ve sus cursos asignados"""
    # Buscar docente por email
    docente = db.query(Docente).filter(Docente.email == current_user.email).first()
    if not docente and current_user.rol != "admin":
        raise HTTPException(404, "Perfil docente no encontrado")
    
    if current_user.rol == "admin":
        cursos = db.query(CursoDB).all()
    else:
        cursos = db.query(CursoDB).filter(CursoDB.docente_id == docente.id).all()
    
    return {
        "docente": f"{current_user.nombre} {current_user.apellido}",
        "total_cursos": len(cursos),
        "cursos": [
            {
                "id": c.id,
                "codigo": c.codigo,
                "nombre": c.nombre,
                "semestre": c.semestre,
                "horas_semanales": c.horas_semanales
            }
            for c in cursos
        ]
    }



# ========== ADMIN - GESTIÓN DE USUARIOS ==========

@router.get("/admin/usuarios")
async def listar_usuarios_admin(
    current_user: Usuario = Depends(require_role(["admin"])),
    db: Session = Depends(get_db)
):
    """Lista todos los usuarios (solo admin)"""
    usuarios = db.query(Usuario).all()
    return [
        {
            "id": u.id,
            "email": u.email,
            "nombre": u.nombre,
            "apellido": u.apellido,
            "rol": u.rol,
            "activo": u.activo,
            "created_at": u.created_at.isoformat() if u.created_at else None
        }
        for u in usuarios
    ]


@router.get("/usuarios")
async def listar_usuarios_simple(
    rol: Optional[str] = None,
    current_user: Usuario = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Lista usuarios (filtrable por rol) - requiere autenticación"""
    query = db.query(Usuario)
    if rol:
        query = query.filter(Usuario.rol == rol)
    usuarios = query.all()
    return [
        {
            "id": u.id,
            "email": u.email,
            "nombre": u.nombre,
            "apellido": u.apellido,
            "rol": u.rol,
            "activo": u.activo
        }
        for u in usuarios
    ]


@router.put("/admin/usuarios/{usuario_id}")
async def actualizar_usuario_admin(
    usuario_id: int,
    usuario_data: dict,
    current_user: Usuario = Depends(require_role(["admin"])),
    db: Session = Depends(get_db)
):
    """Actualiza un usuario (solo admin)"""
    usuario = db.query(Usuario).filter(Usuario.id == usuario_id).first()
    if not usuario:
        raise HTTPException(404, "Usuario no encontrado")

    if "nombre" in usuario_data:
        usuario.nombre = usuario_data["nombre"]
    if "apellido" in usuario_data:
        usuario.apellido = usuario_data["apellido"]
    if "email" in usuario_data:
        usuario.email = usuario_data["email"]
    if "rol" in usuario_data:
        usuario.rol = usuario_data["rol"]
    if "activo" in usuario_data:
        usuario.activo = usuario_data["activo"]
    if "password" in usuario_data and usuario_data["password"]:
        usuario.hashed_password = get_password_hash(usuario_data["password"])

    db.commit()
    db.refresh(usuario)
    return {
        "id": usuario.id,
        "email": usuario.email,
        "nombre": usuario.nombre,
        "apellido": usuario.apellido,
        "rol": usuario.rol,
        "activo": usuario.activo
    }


@router.delete("/admin/usuarios/{usuario_id}")
async def eliminar_usuario_admin(
    usuario_id: int,
    current_user: Usuario = Depends(require_role(["admin"])),
    db: Session = Depends(get_db)
):
    """Elimina un usuario (solo admin)"""
    usuario = db.query(Usuario).filter(Usuario.id == usuario_id).first()
    if not usuario:
        raise HTTPException(404, "Usuario no encontrado")

    if usuario.id == current_user.id:
        raise HTTPException(400, "No puedes eliminar tu propio usuario")

    db.delete(usuario)
    db.commit()
    return {"mensaje": "Usuario eliminado correctamente"}

@router.post("/generar-horarios-avanzado")
def generar_horarios_avanzado(
    request: HorariosAvanzadoRequest,
    current_user: Usuario = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    Genera múltiples alternativas de horario para el semestre indicado.
    Cada alternativa usa una estrategia distinta (mañana, tarde, noche,
    balanceado o aleatoria) y se ordena por puntuación descendente.
    """
    # --- Cargar datos desde BD ---
    cursos = db.query(CursoDB).filter(CursoDB.semestre == request.semestre).all()
    if not cursos:
        raise HTTPException(
            status_code=404,
            detail=f"No se encontraron cursos para el semestre {request.semestre}"
        )
 
    aulas   = db.query(Aula).filter(Aula.activa == True).all()
    docentes = db.query(Docente).filter(Docente.activo == True).all()
 
    if not aulas:
        raise HTTPException(status_code=422, detail="No hay aulas activas configuradas")
 
    # --- Serializar a dict para el motor (evita dependencias de SQLAlchemy) ---
    def curso_to_dict(c) -> dict:
        return {
            "id":              c.id,
            "codigo":          c.codigo,
            "nombre":          c.nombre,
            "creditos":        c.creditos,
            "semestre":        c.semestre,
            "tipo":            c.tipo or "teoria",
            "horas_semanales": c.horas_semanales or 4,
            "max_estudiantes": c.max_estudiantes or 30,
            "docente_id":      c.docente_id,
            "programa_id":     c.programa_id,
        }
 
    def aula_to_dict(a) -> dict:
        return {
            "id":         a.id,
            "codigo":     a.codigo,
            "nombre":     a.nombre,
            "capacidad":  a.capacidad or 30,
            "tipo":       a.tipo or "teoria",
        }
 
    def docente_to_dict(d) -> dict:
        return {
            "id":     d.id,
            "codigo": d.codigo,
            "nombre": f"{d.nombre} {d.apellido}",
        }
 
    cursos_data   = [curso_to_dict(c)   for c in cursos]
    aulas_data    = [aula_to_dict(a)    for a in aulas]
    docentes_data = [docente_to_dict(d) for d in docentes]
 
    # --- Ejecutar motor CSP ---
    engine = CSPEngineProfesional()
    resultado = engine.generar_multiples_horarios(
        cursos=cursos_data,
        aulas=aulas_data,
        docentes=docentes_data,
        semestre=request.semestre,
        num_soluciones=request.num_soluciones,
        preferencias_estudiante={"turno_preferido": request.turno_preferido},
    )
 
    if not resultado["soluciones"]:
        raise HTTPException(
            status_code=422,
            detail="No se pudieron generar horarios. Verifica que hay docentes y aulas suficientes."
        )
 
    logger.info(
        f"[generar-horarios-avanzado] usuario={current_user.email} "
        f"semestre={request.semestre} soluciones={resultado['total_soluciones']}"
    )
 
    return resultado


# ========== INCLUIR ROUTER ==========
app.include_router(router)

# ========== HEALTH CHECK ==========
@app.get("/health")
def health_check():
    return {"status": "ok", "timestamp": datetime.now().isoformat()}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000, reload=True)